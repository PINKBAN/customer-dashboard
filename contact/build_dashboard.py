#!/usr/bin/env python3
"""
客户联络看板 — 一键更新脚本
用法: python build_dashboard.py [销售数据.xlsx] [客户联络.xlsx]
      python build_dashboard.py --no-upload  # 仅生成本地HTML，不上传Supabase

流程:
  1. 加载 Excel 数据
  2. 计算拿货周期
  3. 生成 FULL_DATA
  4. 上传到 Supabase (data_snapshots 表)
  5. 更新 index.html (以 Supabase 为主的轻量版)
  6. 生成 客户联络看板.html (内嵌数据的离线备用版)
"""

import sys, os, json, math, re
from datetime import datetime, date, timedelta
from collections import defaultdict
import numpy as np

# ---- 尝试导入 openpyxl ----
try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# ---- 配置 ----
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_HTML = os.path.join(ROOT, "index.html")
OUTPUT_HTML = os.path.join(ROOT, "index.html")
OUTPUT_HTML2 = os.path.join(ROOT, "contact", "客户联络看板.html")
SUPABASE_URL = "https://rmuhugjufmoghzyynvrb.supabase.co"
SUPABASE_KEY = "sb_publishable_sv9TseWVsQ6dhkqnPYbiqw_aW1Df-5D"

# ---- 工具 ----
def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if not val:
        return None
    s = str(val).strip()
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"]:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None

# ---- 加载数据 ----
def load_customers(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    customers = []
    for row in rows:
        if not row[0]:
            continue
        contact = str(row[10] or "").strip()
        tags = str(row[16] or "").strip()
        tags = ",".join(t for t in tags.replace("，", ",").split(",") if t.strip())
        customers.append({
            "name": str(row[0]).strip(),
            "manager": str(row[1] or "").strip(),
            "lastTradeDate": parse_date(row[3]) or parse_date(row[2]),
            "lastAmount": float(row[5] or 0),
            "receivable": float(row[6] or 0),
            "yearAmount": float(row[7] or 0),
            "monthAmount": float(row[8] or 0),
            "lastMonthAmount": float(row[11] or 0),
            "contact": contact,
            "level": int(row[13] or 0),
            "payment": float(row[15] or 0),
            "tags": tags,
            "noTradeDays": int(row[4] or 0),
        })
    wb.close()
    return customers

def load_sales(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    # 读取表头，自动识别列位置
    header = [str(c.value or "").strip() for c in ws[1]]
    date_col = customer_col = brand_col = None
    # 日期列优先级：单据日期 > 出库日期 > 带"日期"的列
    for keyword in ("单据日期", "出库日期"):
        for i, h in enumerate(header):
            if h == keyword:
                date_col = i
                break
        if date_col is not None:
            break
    if date_col is None:
        for i, h in enumerate(header):
            if "日期" in h:
                date_col = i
                break
    for i, h in enumerate(header):
        if customer_col is None and h == "客户":
            customer_col = i
        if brand_col is None and ("产品品牌" in h or "商品分类" in h):
            brand_col = i
    # 兜底：用老格式默认值
    if date_col is None:
        date_col = 0
    if customer_col is None:
        customer_col = 1
    print(f"  销售表列识别: 日期=col{date_col}, 客户=col{customer_col}, 品牌=col{brand_col}")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    customer_dates = defaultdict(list)
    customer_brands = defaultdict(lambda: defaultdict(list))
    for row in rows:
        name = str(row[customer_col] or "").strip() if customer_col < len(row) else ""
        if not name:
            continue
        d = parse_date(row[date_col]) if date_col < len(row) else None
        if d:
            customer_dates[name].append(d)
            if brand_col is not None and brand_col < len(row):
                brand = str(row[brand_col] or "").strip()
                if brand:
                    customer_brands[name][brand].append(d)
    wb.close()
    return customer_dates, customer_brands

def analyze_product_cycles(brand_dates, today):
    results = []
    for brand, dates in brand_dates.items():
        if len(dates) < 2:
            continue
        unique_dates = sorted(set(dates))
        intervals = []
        for i in range(1, len(unique_dates)):
            delta = (unique_dates[i] - unique_dates[i-1]).days
            if delta > 0:
                intervals.append(delta)
        if not intervals:
            continue
        intervals_arr = np.array(intervals)
        median_cycle = round(float(np.median(intervals_arr)), 1)
        mean_cycle = float(np.mean(intervals_arr))
        cv = float(np.std(intervals_arr)) / mean_cycle if mean_cycle > 0 else 0
        regularity = round(max(0, min(1, 1 - cv)), 2)
        last_date = unique_dates[-1]
        results.append({
            "brand": brand,
            "cycleDays": median_cycle,
            "regularity": regularity,
            "lastDate": last_date.strftime("%Y-%m-%d"),
            "count": len(dates)
        })
    results.sort(key=lambda x: x["count"], reverse=True)
    return results[:5]

# ---- 周期分析 ----
def analyze_cycle(dates, today):
    if len(dates) < 2:
        return {
            "hasCycle": False, "cycleDays": 0, "regularity": 0,
            "predictedNext": "", "daysToNext": 0,
            "cycleLastDate": "", "cycleUrgency": "normal",
            "suggestion": "近期有交易，暂无周期数据"
        }
    unique_dates = sorted(set(dates))
    if len(unique_dates) < 2:
        return {
            "hasCycle": False, "cycleDays": 0, "regularity": 0,
            "predictedNext": "", "daysToNext": 0,
            "cycleLastDate": "", "cycleUrgency": "normal",
            "suggestion": "近期有交易，暂无周期数据"
        }
    intervals = []
    for i in range(1, len(unique_dates)):
        delta = (unique_dates[i] - unique_dates[i-1]).days
        if delta > 0:
            intervals.append(delta)
    if not intervals:
        return {
            "hasCycle": False, "cycleDays": 0, "regularity": 0,
            "predictedNext": "", "daysToNext": 0,
            "cycleLastDate": "", "cycleUrgency": "normal",
            "suggestion": "近期有交易，暂无周期数据"
        }
    intervals_arr = np.array(intervals)
    median_cycle = round(float(np.median(intervals_arr)), 1)
    mean_cycle = float(np.mean(intervals_arr))
    if mean_cycle > 0:
        cv = float(np.std(intervals_arr)) / mean_cycle
        regularity = round(max(0, min(1, 1 - cv)), 2)
    else:
        regularity = 0
    last_date = unique_dates[-1]
    predicted_next = last_date + timedelta(days=int(median_cycle))
    days_to_next = (predicted_next - today).days
    if days_to_next < -14:
        urgency = "urgent"
    elif days_to_next < 0:
        urgency = "overdue"
    elif days_to_next <= 3:
        urgency = "soon"
    elif days_to_next <= 7:
        urgency = "upcoming"
    else:
        urgency = "normal"
    if urgency == "urgent":
        suggestion = f"已超期{abs(days_to_next)}天，需立即联系"
    elif urgency == "overdue":
        suggestion = f"刚超期{abs(days_to_next)}天，建议本周联系"
    elif urgency == "soon":
        suggestion = f"预计{days_to_next}天后到拿货周期，建议提前联系"
    elif urgency == "upcoming":
        suggestion = f"距拿货周期还有{days_to_next}天"
    else:
        suggestion = f"距拿货周期还有{days_to_next}天"
    return {
        "hasCycle": True, "cycleDays": median_cycle,
        "regularity": regularity,
        "predictedNext": predicted_next.strftime("%Y-%m-%d"),
        "daysToNext": days_to_next,
        "cycleLastDate": last_date.strftime("%Y-%m-%d"),
        "cycleUrgency": urgency, "suggestion": suggestion
    }

# ---- 匹配客户 ----
def match_customers(customer_list, sales_dict, brand_dict):
    name_to_dates = {}
    name_to_brands = {}
    for cname, dates in sales_dict.items():
        name_to_dates[cname] = dates
        name_to_brands[cname] = dict(brand_dict.get(cname, {}))
    for c in customer_list:
        cname = c["name"]
        if cname in name_to_dates and len(name_to_dates[cname]) >= 2:
            c["_dates"] = name_to_dates[cname]
            c["_brands"] = name_to_brands.get(cname, {})
            continue
        found = False
        for sname, dates in name_to_dates.items():
            if cname in sname or sname in cname:
                if len(dates) >= 2 and not found:
                    c["_dates"] = dates
                    c["_brands"] = name_to_brands.get(sname, {})
                    found = True
        if not found:
            c["_dates"] = name_to_dates.get(cname, [])
            c["_brands"] = name_to_brands.get(cname, {})

def build_full_data(customer_list, today):
    result = []
    for i, c in enumerate(customer_list):
        cycle = analyze_cycle(c.get("_dates", []), today)
        brands = c.get("_brands", {})
        product_cycles = analyze_product_cycles(brands, today) if brands else []
        item = {
            "id": i,
            "name": c["name"],
            "manager": c["manager"],
            "lastTradeDate": c["lastTradeDate"].strftime("%Y-%m-%d") if c["lastTradeDate"] else "",
            "noTradeDays": c["noTradeDays"],
            "lastAmount": c["lastAmount"],
            "receivable": c["receivable"],
            "yearAmount": c["yearAmount"],
            "monthAmount": c["monthAmount"],
            "lastMonthAmount": c["lastMonthAmount"],
            "contact": c["contact"],
            "level": c["level"],
            "payment": c["payment"],
            "tags": c["tags"],
            "productSummary": product_cycles,
            **cycle
        }
        result.append(item)
    return result

# ---- 生成 HTML ----
def generate_html(full_data, template_path, today):
    with open(template_path, "r", encoding="utf-8") as f:
        html = f.read()
    date_str = today.strftime("%Y-%m-%d")
    html = re.sub(
        r'数据截止：\d{4}-\d{2}-\d{2}',
        f'数据截止：{date_str}',
        html
    )
    json_str = json.dumps(full_data, ensure_ascii=False, separators=(",", ": "))
    new_line = f"var FULL_DATA = {json_str};"
    pattern = r'var FULL_DATA = \[.*?\];'
    html = re.sub(pattern, new_line.replace('\\', '\\\\'), html, count=1)
    return html

# ---- Supabase 上传 ----
def upload_to_supabase(full_data, source_file):
    """上传 FULL_DATA 到 Supabase data_snapshots 表"""
    import urllib.request
    import urllib.error

    api_url = f"{SUPABASE_URL}/rest/v1/data_snapshots"

    # 1) 将旧快照设为 inactive
    deactivate_req = urllib.request.Request(
        f"{api_url}?is_active=eq.true",
        data=json.dumps({"is_active": False}).encode('utf-8'),
        method='PATCH'
    )
    deactivate_req.add_header('apikey', SUPABASE_KEY)
    deactivate_req.add_header('Authorization', f'Bearer {SUPABASE_KEY}')
    deactivate_req.add_header('Content-Type', 'application/json')
    deactivate_req.add_header('Prefer', 'return=minimal')
    try:
        urllib.request.urlopen(deactivate_req)
    except Exception as e:
        print(f"  [警告] 停用旧快照失败: {e}")

    # 2) 插入新快照
    payload = json.dumps({
        "data": full_data,
        "customer_count": len(full_data),
        "is_active": True,
        "source_file": source_file
    }, ensure_ascii=False).encode('utf-8')

    req = urllib.request.Request(api_url, data=payload, method='POST')
    req.add_header('apikey', SUPABASE_KEY)
    req.add_header('Authorization', f'Bearer {SUPABASE_KEY}')
    req.add_header('Content-Type', 'application/json')
    req.add_header('Prefer', 'return=minimal')

    try:
        urllib.request.urlopen(req)
        return True
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace') if e.fp else ''
        print(f"  Supabase HTTP {e.code}: {body[:200]}")
        return False
    except Exception as e:
        print(f"  Supabase 连接失败: {e}")
        return False

# ---- 主流程 ----
def main():
    today = date.today()
    no_upload = '--no-upload' in sys.argv

    print(f"=== 客户联络看板 — 数据更新 ===")
    print(f"日期: {today}")

    # 找到客户联络文件
    contact_dir = os.path.join(ROOT, "contact")
    contact_file = os.path.join(contact_dir, "客户联络.xlsx")

    # 自动检测最新版本的客户联络文件（18列，小文件）
    if os.path.isdir(contact_dir):
        candidates = []
        for f in os.listdir(contact_dir):
            fpath = os.path.join(contact_dir, f)
            if not f.endswith('.xlsx') or f == '客户联络.xlsx':
                continue
            if os.path.getsize(fpath) >= 500_000:
                continue
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = wb.active
                if ws.max_column == 18 and ws.max_row < 2000:
                    candidates.append((os.path.getmtime(fpath), fpath))
                wb.close()
            except Exception:
                pass
        if candidates:
            candidates.sort(reverse=True)
            contact_file = candidates[0][1]
            print(f"自动选择最新客户联络: {os.path.basename(contact_file)}")

    # 找到销售数据文件（大文件 >500KB，且不含"客户联络"）
    if len(sys.argv) > 1 and sys.argv[1] not in ('--no-upload',):
        sales_file = sys.argv[1]
        if not os.path.isabs(sales_file):
            sales_file = os.path.join(contact_dir, sales_file)
    else:
        xlsx_files = []
        if os.path.isdir(contact_dir):
            for f in os.listdir(contact_dir):
                if f.endswith('.xlsx') and '客户联络' not in f:
                    fpath = os.path.join(contact_dir, f)
                    if os.path.getsize(fpath) > 500_000:
                        xlsx_files.append((os.path.getmtime(fpath), fpath))
        if xlsx_files:
            xlsx_files.sort(reverse=True)
            sales_file = xlsx_files[0][1]
            print(f"自动选择最新销售文件: {os.path.basename(sales_file)}")
        else:
            print("错误: 未找到销售数据文件（>500KB的xlsx），请放入 contact 文件夹")
            sys.exit(1)

    print(f"销售数据: {sales_file}")
    print(f"客户联络: {contact_file}")

    # 加载数据
    print("\n[1/4] 加载客户联络表...")
    customers = load_customers(contact_file)
    print(f"  客户数: {len(customers)}")

    inactive = 0
    for c in customers:
        if c["noTradeDays"] >= 100:
            c["manager"] = "未激活客户"
            inactive += 1
    print(f"  未激活客户(>=100天未交易): {inactive}")

    print("[2/4] 加载销售明细...")
    sales, brands = load_sales(sales_file)
    top = sorted(sales.items(), key=lambda x: len(x[1]), reverse=True)[:5]
    brand_count = sum(1 for b in brands.values() if b)
    print(f"  有销售记录客户数: {len(sales)}")
    print(f"  有品牌数据客户数: {brand_count}")
    print(f"  Top5: {[(n, len(d)) for n, d in top]}")

    print("[3/4] 匹配客户 + 计算周期...")
    match_customers(customers, sales, brands)
    full_data = build_full_data(customers, today)

    with_cycle = sum(1 for d in full_data if d["hasCycle"])
    urgent = sum(1 for d in full_data if d["cycleUrgency"] == "urgent")
    overdue = sum(1 for d in full_data if d["cycleUrgency"] == "overdue")
    print(f"  总客户: {len(full_data)}")
    print(f"  有周期数据: {with_cycle}")
    print(f"  紧急(超期>14天): {urgent}")
    print(f"  超期: {overdue}")

    print("[4/4] 生成 HTML + 上传 Supabase...")

    if not os.path.exists(TEMPLATE_HTML):
        print(f"  错误: 模板文件不存在 {TEMPLATE_HTML}")
        sys.exit(1)

    # 上传到 Supabase
    source_name = os.path.basename(sales_file)
    if not no_upload:
        ok = upload_to_supabase(full_data, source_name)
        if ok:
            print(f"  -> Supabase 上传成功 ({len(full_data)} 条客户数据)")
        else:
            print("  -> Supabase 上传失败，仅更新本地文件")
    else:
        print("  -> 跳过 Supabase 上传 (--no-upload)")

    # 生成轻量版 index.html（Supabase 为主的版本）
    html = generate_html(full_data, TEMPLATE_HTML, today)
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  -> {OUTPUT_HTML}")

    # 生成离线备用版 客户联络看板.html（内嵌数据）
    html2 = generate_html(full_data, TEMPLATE_HTML, today)
    with open(OUTPUT_HTML2, "w", encoding="utf-8") as f:
        f.write(html2)
    print(f"  -> {OUTPUT_HTML2}")

    print("\n完成!")
    if not no_upload:
        print("  所有用户刷新页面即可看到最新数据（无需重新部署 HTML）")

if __name__ == "__main__":
    main()
